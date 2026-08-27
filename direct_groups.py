# -*- coding: utf-8 -*-
"""Марочные и модельные группы для Яндекс.Директа.

Берёт выгрузку кампании из кабинета, дописывает в неё группы по маркам
и моделям и сохраняет новый файл — его загружают обратно в Директ.

Запуск:  ./venv/bin/python direct_groups.py ~/Desktop/713906077.xlsx
Результат: ~/Desktop/<имя>-с-группами.xlsx

Почему не собираем файл с нуля: набор колонок зависит от версии кабинета,
и угадывать его вслепую — верный способ получить ошибку импорта.
"""
import collections
import json
import os
import re
import sys

import openpyxl

from seo_lib import cat_url

SITE = "https://dvigatel-ekb.ru"
MIN_BRAND = 3      # столько позиций должно быть у марки, иначе страница пустовата
MIN_MODEL = 8      # у модели порог выше: платный трафик на три позиции не окупается
BID = 40           # ставка, ₽ за клик
H_LIMIT, T_LIMIT = 56, 81

# Модели, которых нет в одиночных записях: в прайсе они приходят только
# сдвоенными («CITROEN/PEUGEOT»), и марку по данным не определить.
MODEL_BRAND = {
    "C1": "CITROEN", "C2": "CITROEN", "C3": "CITROEN", "C4": "CITROEN",
    "C5": "CITROEN", "C6": "CITROEN", "C8": "CITROEN", "DS3": "CITROEN",
    "DS4": "CITROEN", "DS5": "CITROEN", "BERLINGO": "CITROEN",
    "XSARA": "CITROEN", "JUMPY": "CITROEN", "JUMPER": "CITROEN",
    "206": "PEUGEOT", "207": "PEUGEOT", "208": "PEUGEOT", "301": "PEUGEOT",
    "306": "PEUGEOT", "307": "PEUGEOT", "308": "PEUGEOT", "406": "PEUGEOT",
    "407": "PEUGEOT", "408": "PEUGEOT", "508": "PEUGEOT", "607": "PEUGEOT",
    "1007": "PEUGEOT", "3008": "PEUGEOT", "4007": "PEUGEOT", "5008": "PEUGEOT",
    "PARTNER": "PEUGEOT", "BOXER": "PEUGEOT", "EXPERT": "PEUGEOT",
}

# латиница внутри синонимов — это коды коробок, а не названия марок
NOT_A_NAME = {"cvt", "dsg", "at", "mt"}


def load():
    return json.load(open("data/parts.json", encoding="utf-8"))["items"]


def brands_of(it):
    return it.get("bs") or [it["b"]]


def models_of(it):
    return it.get("ms") or ([it["m"]] if it["m"] else [])


def brand_ru(items):
    """Русские написания марок — из поисковых синонимов каталога."""
    common = set("двигатель мотор двс акпп автомат автоматическая коробка передач "
                 "вариатор робот dsg мкпп механика механическая раздатка "
                 "раздаточная".split())
    hits = collections.defaultdict(collections.Counter)
    for it in items:
        b = str(it.get("b", ""))
        if "/" in b:
            continue
        for w in str(it.get("x", "")).split():
            if w not in common and w not in NOT_A_NAME and not w.isdigit():
                hits[b.upper()][w] += 1
    return {b: c.most_common(1)[0][0] for b, c in hits.items() if c}


def model_owner(items):
    """Модель → марка, по записям с одной маркой. Для остальных — словарь выше."""
    owner = collections.defaultdict(collections.Counter)
    for it in items:
        b = str(it.get("b", ""))
        if "/" in b:
            continue
        for m in models_of(it):
            owner[m.upper()][b.upper()] += 1
    out = {m: c.most_common(1)[0][0] for m, c in owner.items() if c}
    for m, b in MODEL_BRAND.items():
        out.setdefault(m, b)
    return out


def belongs(model, brand, owners):
    """Пара марка+модель настоящая? «Audi Golf» появляется из сдвоенных записей."""
    for part in re.split(r"[/,]", model):
        part = part.strip().upper()
        if not part:
            continue
        known = owners.get(part)
        if known:
            return known == brand.upper()
    return True     # модель нигде не встречалась одна — оставляем, вреда нет


def money(n):
    return f"{int(n):,}".replace(",", " ")


def plural(n, one, few, many):
    n10, n100 = n % 10, n % 100
    if n10 == 1 and n100 != 11:
        return one
    if 2 <= n10 <= 4 and not 12 <= n100 <= 14:
        return few
    return many


def model_names(model):
    """«3 / Axela» → ['3', 'axela']: люди ищут и так, и так."""
    return [p.strip() for p in re.split(r"[/,]", model) if p.strip()]


def keys_for(cat, ru, lat, model=None):
    """Ключевые фразы. Марка пишется и кириллицей, и латиницей."""
    names = [ru, lat.lower()]
    out = []
    if cat == "двс":
        for n in names:
            tail = f"{n} {model}".strip() if model else n
            out += [f"контрактный двигатель {tail}", f"двигатель {tail} бу",
                    f"купить двигатель {tail}"]
        if not model:
            out += [f"контрактный двс {ru}", f"двигатель {ru} екатеринбург"]
    else:
        for n in names:
            tail = f"{n} {model}".strip() if model else n
            out += [f"контрактная акпп {tail}", f"акпп {tail} бу",
                    f"коробка автомат {tail} бу"]
        if not model:
            out += [f"вариатор {ru} бу", f"акпп {ru} екатеринбург"]
    seen, uniq = set(), []
    for k in out:
        k = re.sub(r"\s+", " ", k).strip().lower()
        if k not in seen:
            seen.add(k)
            uniq.append(k)
    return uniq


def texts_for(cat, lat, cnt, price, model=None):
    """Заголовки и тексты. Всё, что длиннее лимита, отсеется проверкой ниже."""
    what = f"{lat.title()} {model}".strip() if model else lat.title()
    if cat == "двс":
        heads = [f"Контрактные двигатели {what}",
                 f"{cnt} в наличии, Екатеринбург",
                 f"Двигатели {what} от {money(price)} ₽",
                 "Подберём по VIN бесплатно",
                 "Пробег и фото каждого агрегата",
                 "Отправим в любой город России",
                 "Установка в своём сервисе"]
        body = [f"{cnt} {plural(cnt, 'двигатель', 'двигателя', 'двигателей')} "
                f"в наличии. Пробег, номер и OEM в карточке.",
                "Екатеринбург, свой сервис. Привезём и поставим, гарантия на работу.",
                "Фото и видео агрегата до оплаты. Отправка транспортной компанией."]
    else:
        heads = [f"Контрактные АКПП {what}",
                 f"{cnt} в наличии, Екатеринбург",
                 f"АКПП {what} от {money(price)} ₽",
                 "Подберём коробку по VIN бесплатно",
                 "Код коробки и пробег в карточке",
                 "Отправим в любой город России",
                 "Установка в своём сервисе"]
        body = [f"{cnt} {plural(cnt, 'коробка', 'коробки', 'коробок')} "
                f"в наличии. Пробег и код коробки в карточке.",
                "Екатеринбург, свой сервис. Привезём и поставим, гарантия на работу.",
                "Фото и видео агрегата до оплаты. Отправка транспортной компанией."]
    heads = [h for h in heads if len(h) <= H_LIMIT][:7]
    body = [t for t in body if len(t) <= T_LIMIT][:3]
    return heads, body


def collect():
    """Все группы, которые имеет смысл откручивать."""
    items = load()
    ru_map, owners = brand_ru(items), model_owner(items)

    pairs, triples = collections.defaultdict(list), collections.defaultdict(list)
    for it in items:
        if it["n"] not in ("двс", "акпп"):
            continue
        for b in brands_of(it):
            if "/" in b:            # сдвоенную марку как таковую не рекламируем
                continue
            pairs[(it["n"], b)].append(it)
            for m in models_of(it):
                triples[(it["n"], b, m)].append(it)

    def lowest(lst):
        p = [x["p"] for x in lst if isinstance(x.get("p"), (int, float)) and x["p"] > 0]
        return min(p) if p else 0

    groups = []
    for (cat, brand), lst in sorted(pairs.items(), key=lambda kv: -len(kv[1])):
        if len(lst) < MIN_BRAND or not lowest(lst):
            continue
        ru = ru_map.get(brand.upper())
        if not ru:
            continue
        groups.append(dict(cat=cat, brand=brand, model=None, ru=ru, lat=brand,
                           cnt=len(lst), price=lowest(lst),
                           url=cat_url(cat, brand), level="марка"))

    for (cat, brand, model), lst in sorted(triples.items(), key=lambda kv: -len(kv[1])):
        if len(lst) < MIN_MODEL or not lowest(lst):
            continue
        ru = ru_map.get(brand.upper())
        if not ru or not belongs(model, brand, owners):
            continue
        for name in model_names(model)[:1]:     # первое написание — основное
            groups.append(dict(cat=cat, brand=brand, model=name, ru=ru, lat=brand,
                               cnt=len(lst), price=lowest(lst),
                               url=cat_url(cat, brand, model), level="модель"))
    return groups


def campaign_tag(g):
    base = "dvs" if g["cat"] == "двс" else "akpp"
    slug = re.sub(r"[^a-z0-9]+", "_", (g["lat"] + ("_" + g["model"] if g["model"] else "")).lower())
    return f"{base}_{slug}".strip("_")[:60]


def main(src):
    groups = collect()
    wb = openpyxl.load_workbook(src)
    ws = wb["Тексты"]

    sample = {"двс": 12, "акпп": 21}            # строки-образцы из выгрузки
    row = ws.max_row + 1
    num = max(int(ws.cell(r, 5).value or 0) for r in range(12, ws.max_row + 1))

    made = 0
    for g in groups:
        src_row = sample[g["cat"]]
        heads, body = texts_for(g["cat"], g["lat"], g["cnt"], g["price"], g["model"])
        keys = keys_for(g["cat"], g["ru"], g["lat"], g["model"])
        num += 1
        title = f'{"Двигатели" if g["cat"] == "двс" else "АКПП"} {g["lat"].title()}'
        if g["model"]:
            title += " " + g["model"]
        link = f'{SITE}/{g["url"]}?utm_source=direct&utm_medium=cpc&utm_campaign={campaign_tag(g)}'

        for key in keys + ["---autotargeting"]:
            ws.cell(row, 1, "-")
            ws.cell(row, 2, "Комбинаторное")
            ws.cell(row, 4, title[:255])
            ws.cell(row, 5, num)
            ws.cell(row, 7, key)
            for i, h in enumerate(heads):
                ws.cell(row, 15 + i, h)
            for i, t in enumerate(body):
                ws.cell(row, 22 + i, t)
            for col in (35, 36, 47, 49, 51, 53, 54, 65):
                ws.cell(row, col, ws.cell(src_row, col).value)
            ws.cell(row, 47, link)
            ws.cell(row, 51, BID)
            row += 1
        made += 1

    out = os.path.splitext(src)[0] + "-с-группами.xlsx"
    wb.save(out)

    by_level = collections.Counter(g["level"] for g in groups)
    print(f"групп добавлено: {made}  (марок {by_level['марка']}, моделей {by_level['модель']})")
    print(f"строк в файле: {row - 1}")
    print(f"файл: {out}")


if __name__ == "__main__":
    main(os.path.expanduser(sys.argv[1] if len(sys.argv) > 1
                            else "~/Desktop/713906077.xlsx"))
